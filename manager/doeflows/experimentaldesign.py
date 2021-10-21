# 
# experimentaldesign.py
# 
# Author(s):
# Matteo Spallanzani <spmatteo@iis.ee.ethz.ch>
# 
# Copyright (c) 2020-2021 ETH Zurich.
# 
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
# 
# http://www.apache.org/licenses/LICENSE-2.0
# 
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
#


import collections.abc
from collections import namedtuple
from enum import EnumMeta, IntEnum
import importlib
import inspect
import copy
import os
import shutil
import json
import pandas as pd
from openpyxl import load_workbook

from typing import Tuple, List, Dict
from typing import Union


_CFG_FILE_EXTENSION = 'json'
_EDL_FILE_EXTENSION = 'xlsx'  # degrees of freedom and setups of the experimental design will be logged in an Excel workbook


def patch_dictionary(config: Dict, patch: Dict):
    """Traverse and update dictionary, recursively."""
    # https://stackoverflow.com/a/3233356
    for k, v in patch.items():

        if isinstance(v, collections.abc.Mapping):
            config[k] = patch_dictionary(config.get(k, {}), v)
        else:
            config[k] = v

    return config


ExperimentalSetup = namedtuple('ExperimentalSetup', ['dofs_values', 'config_patch'])


class ExperimentalDesign(object):

    def __init__(self, dofs: Tuple[EnumMeta]) -> None:
        """QuantLab assumes that experimental designs are specific to a given
        system package (problem- and topology-sub-package combination).
        According to this criterion, we decided to automate the management of
        experimental designs by imposing the following structure. For those
        systems packages where a user would like to run large experimental
        designs (i.e., designs consisting of a large number of experimental
        units), it is necessary to create a `doe` sub-package. This package
        will contain pairs of files, each pair describing a specific
        experimental design. To distinguish experimental designs, they should
        be given different names. If we suppose that the experimental design
        is called [EXPDESIGN], the files will be the following:

        * [expdesign].py, a Python module implementing a sub-class of
          `ExperimentalDesign` declining the `_generate_protoconfigs`
          method;
        * [expdesign].json a JSON file containing a "basic" QuantLab
          configuration file that will be used as the default configuration
          of the system on top of which patches changing its hyper-parameters
          will be applied.

        Note that we establish the convention of naming the experimental
        design using the UPPERCASE or CapitalisedCase conventions, whereas the
        letters appearing in the filenames should be lowercase.

        """

        self._dofs        = dofs  # degrees of freedom
        self._base_config = self._load_base_config()
        self._setups      = self._generate_experimental_setups()

    @property
    def dofs(self) -> Tuple[EnumMeta]:
        return self._dofs

    def _load_base_config(self) -> Dict:
        """Load the basic configuration dictionary that will be patched."""
        class_ = self.__class__
        exp_design_folder = os.path.dirname(inspect.getfile(class_))
        cfg_file_filename = '.'.join([class_.__name__.upper(), _CFG_FILE_EXTENSION])

        with open(os.path.join(exp_design_folder, cfg_file_filename), 'r') as fp:
            base_config = json.load(fp)

        return base_config

    @property
    def base_config(self) -> Dict:
        return self._base_config

    def _generate_experimental_setups(self) -> List[ExperimentalSetup]:
        """Generate configurations and append them to the experiments list."""
        raise NotImplementedError

    @property
    def setups(self) -> List[ExperimentalSetup]:
        return self._setups

    def patch_base_config(self, setup: ExperimentalSetup) -> Dict:
        return patch_dictionary(copy.deepcopy(self._base_config), setup.config_patch)


class ExperimentalUnitStatus(IntEnum):
    CONFIGURED = 0
    PROCESSING = 1
    COMPLETED  = 2


class ExperimentalDesignLogger(object):

    def __init__(self,
                 problem:    str,
                 topology:   str,
                 exp_design: str) -> None:

        self._problem    = problem
        self._topology   = topology
        self._exp_design = exp_design

        self.ed           = self._load_exp_design()
        self._path_edl    = self._get_path_edl()
        self._eu_register = None  # record of configured experimental units

    def _load_exp_design(self) -> ExperimentalDesign:
        system_package_path = '.'.join(['systems', self._problem, self._topology])
        doe_module          = importlib.import_module('.doe', package=system_package_path)
        return getattr(doe_module, self._exp_design)()

    def _get_path_edl(self) -> str:
        # I assume that this object will live in a session of the Python
        # interpreter that has been launched from QuantLab's home.
        path_logs    = os.path.join('systems', self._problem, self._topology, 'logs')
        edl_filename = '.'.join([self._exp_design, _EDL_FILE_EXTENSION])
        return os.path.join(path_logs, edl_filename)

    def write_args(self, args) -> None:

        with pd.ExcelWriter(self._path_edl, engine='openpyxl') as ew:

            args_mapping = [(k, v) for k, v in vars(args).items() if k not in ['doeflows', 'func', 'horovod']]
            df = pd.DataFrame(data=args_mapping, columns=['key', 'value'])
            df.to_excel(ew, sheet_name='CommandLine', index=False)

    def write_dofs(self) -> None:

        assert os.path.isfile(self._path_edl)
        workbook = load_workbook(self._path_edl)

        with pd.ExcelWriter(self._path_edl, engine='openpyxl') as ew:

            ew.book   = workbook
            ew.sheets = {worksheet.title: worksheet for worksheet in workbook.worksheets}

            dofs = {dof.__name__: [(v.name, v.value) for v in dof] for dof in self.ed.dofs}
            for name, mapping in dofs.items():
                df = pd.DataFrame(data=mapping, columns=['name', 'value'])
                df.to_excel(ew, sheet_name=name, index=False)

            ew.save()

    @property
    def eu_register(self) -> List[Tuple]:
        return self._eu_register

    def load_register(self) -> None:

        try:
            df = pd.read_excel(self._path_edl, sheet_name='ExperimentalUnits')
            self._eu_register = [tuple(row) for _, row in df.iterrows()]
        except ValueError or FileNotFoundError:
            self._eu_register = list()  # create empty register

    def update_register(self, status: ExperimentalUnitStatus, id_: int, setup: ExperimentalSetup) -> None:
        # side effect: changes the content of the register of experimental units
        self._eu_register.append((status, id_, *setup.dofs_values))

    def write_register(self) -> None:

        # Panda's `ExcelWriter` does not have a native "append sheet" mode.
        # I found a workaround here:
        #
        #   https://soulsinporto.medium.com/how-to-add-new-worksheets-to-excel-workbooks-with-pandas-47122704fb75
        #
        assert os.path.exists(self._path_edl)
        workbook = load_workbook(self._path_edl)

        with pd.ExcelWriter(self._path_edl, engine='openpyxl') as ew:

            ew.book   = workbook
            ew.sheets = {worksheet.title: worksheet for worksheet in workbook.worksheets}

            df = pd.DataFrame(data=self._eu_register, columns=['Status', 'ID'] + [dof.__name__ for dof in self.ed.dofs])
            df.to_excel(ew, sheet_name='ExperimentalUnits', index=False)

            ew.save()

    def delete_edl(self) -> None:
        assert os.path.isfile(self._path_edl)
        os.remove(self._path_edl)

    def move_edl(self, dst: Union[str, os.PathLike]) -> None:
        shutil.copyfile(self._path_edl, os.path.join(dst, os.path.basename(self._path_edl)))

    def get_path_archive(self) -> str:

        path_logs    = os.path.dirname(self._path_edl)
        path_archive = os.path.join(path_logs, self._exp_design.lower())
        if not os.path.isdir(path_archive):
            os.makedirs(path_archive, exist_ok=True)

        return path_archive

    def is_experiment_archivable(self) -> bool:
        # return all([eu_status == ExperimentalUnitStatus.COMPLETED for (eu_status, eu_id, *eu_dofs_values) in self._eu_register])  # TODO: this should be reworked once the DoE training flow will be completed
        return True
